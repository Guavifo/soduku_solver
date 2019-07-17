import openpyxl
import sys

if len(sys.argv) == 2:
    puzzle_file = sys.argv[1]
else:    
    print('Please specify the Excel file which contains the Soduku puzzle')
    sys.exit()

if puzzle_file[-5:] == '.xlsx':
    print('Checking Excel file...')
else:
    print('Please specify a valid Excel file including the extention')
    sys.exit()

# Load the worksheet into memory
puzzle_workbook = openpyxl.load_workbook(puzzle_file)
ws = puzzle_workbook.active

original = []
line = []

# Read the original puzzle from the Excel file:
for a in range (0, 9):
    for b in range (0, 9):
        line.append(ws.cell(row = a + 2, column = b + 2).value)
    original.append(line)
    line = []

print(original)
solution = original
improved = True

# Start over from this point. Always update the solution matrix.
while improved == True:
    improved = False
    
    # Set up the array of possiblities
    options = [[0 for j in range(9)] for i in range(9)]
    for a in range (0, 9):
        for b in range (0, 9):
            options[a][b] = [1,2,3,4,5,6,7,8,9]
            
    # Remove anything that comes from the solution in memory
    for a in range (0, 9):
        for b in range (0, 9):
            if solution[a][b] != None:
                options[a][b] = [solution[a][b]]
    
    # Show the list of options before getting started
    print('Start of round')
    print(options)
    print()
        # Later, just print the number of possibilities in the matrix - 81
        # The number minus 81 is the number of unknowns remaining
    
    # Eliminate possibilities for overlapping columns and rows
    for a in range (0, 9):
        for b in range (0, 9):
            if solution[a][b] != None:
                target = solution[a][b]
                for c in range (0, 9):
                    if target in options[c][b] and a != c:
                        options[c][b].remove(target)   
                    if target in options[a][c] and b != c:
                        options[a][c].remove(target)

    print('After Lines')
    print(options)
    print()

                    
    #print(options)
                        
    # Check each 3x3 square for options
    # Create 9 lists of 9 tuples, showing what the 3x3 squares are.
    squares = [[(0,0),(0,1),(0,2),(1,0),(1,1),(1,2),(2,0),(2,1),(2,2)],
               [(0,3),(0,4),(0,5),(1,3),(1,4),(1,5),(2,3),(2,4),(2,5)],
               [(0,6),(0,7),(0,8),(1,6),(1,7),(1,8),(2,6),(2,7),(2,8)],
               [(3,0),(3,1),(3,2),(4,0),(4,1),(4,2),(5,0),(5,1),(5,2)],
               [(3,3),(3,4),(3,5),(4,3),(4,4),(4,5),(5,3),(5,4),(5,5)],
               [(3,6),(3,7),(3,8),(4,6),(4,7),(4,8),(5,6),(5,7),(5,8)],
               [(6,0),(6,1),(6,2),(7,0),(7,1),(7,2),(8,0),(8,1),(8,2)],
               [(6,3),(6,4),(6,5),(7,3),(7,4),(7,5),(8,3),(8,4),(8,5)],
               [(6,6),(6,7),(6,8),(7,6),(7,7),(7,8),(8,6),(8,7),(8,8)]]
    print('Squares Taken')
    for square in squares:
        taken = []
        for cell in square:
            a, b = cell
            if len(options[a][b]) == 1:
                taken.append(options[a][b])
        print(taken)
        # Look for these single ones in the rest of the square
        for cell in square:
            a, b = cell
            print(a, b)
            for target in taken:
                print(target[0])
                if (target[0] in options[a][b]) and (len(options[a][b]) > 1):
                    print('Found a target in ', a, b)
                    options[a][b].remove(target[0])

    # Look for any unique options which aren't in the solution matrix
    for a in range (0, 9):
        for b in range (0, 9):
            if len(options[a][b]) == 1 and solution[a][b] is None:
                solution[a][b] = options[a][b][0]
                improved = True
                print('Fixed one')
            # Copy the single value from options up to solution
            # Set improved to true here
    
    # If the solution has changed, go back to step 1
        # if improved = true, it will automatically stay in the loop
    
# Print the puzzle back to the screen for validation
print('Final:')

# Turn this into a stand-alone function
print(solution)
print()

print('+---+---+---+')
for a in range(0, 9):
    print('|', end = '')
    for b in range(0, 9):
        if solution[a][b] == None:
            print('.', end = '')
        else:
            print(solution[a][b], end = '')
        if (b + 1) % 3 == 0:
            print('|', end = '')
    print('')
    if (a + 1) % 3 == 0:
        print('+---+---+---+')

print()

# Define which columns the answer will appear
columns = ['L', 'M', 'N', 'O', 'P', 'Q', 'R', 'S', 'T']

# Write the solution array back to the Excel file:
for a in range(0, 9):
    for b in range(0, 9):
        ws[columns[b] + str(a+2)] = solution[a][b]
    
# Save the file in the current directory
puzzle_workbook.save('Solution_' + puzzle_file)

print('Results saved as Solution_' + puzzle_file)
print()
